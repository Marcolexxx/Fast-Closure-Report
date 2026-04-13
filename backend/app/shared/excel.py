from __future__ import annotations

import logging
from typing import Any, Dict, List, Optional
import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

def _fill_merged_cells(ws: Worksheet) -> None:
    """Populates merged cells in an openpyxl worksheet with the top-left value."""
    # Create a list of merged cell ranges to avoid mutating during iteration
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        top_left_val = ws.cell(row=min_row, column=min_col).value
        
        # Unmerge the range first
        ws.unmerge_cells(str(merged_range))
        
        # Fill all cells in that box
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = top_left_val


def _levenshtein_distance(s1: str, s2: str) -> int:
    """Classic DP Levenshtein distance."""
    if len(s1) < len(s2):
        return _levenshtein_distance(s2, s1)
    if len(s2) == 0:
        return len(s1)
    prev_row = list(range(len(s2) + 1))
    for i, c1 in enumerate(s1):
        curr_row = [i + 1]
        for j, c2 in enumerate(s2):
            cost = 0 if c1 == c2 else 1
            curr_row.append(min(curr_row[j] + 1, prev_row[j + 1] + 1, prev_row[j] + cost))
        prev_row = curr_row
    return prev_row[-1]


def _similarity(a: str, b: str) -> float:
    """Normalized Levenshtein similarity in [0.0, 1.0]."""
    a_low, b_low = a.lower().strip(), b.lower().strip()
    if not a_low or not b_low:
        return 0.0
    max_len = max(len(a_low), len(b_low))
    dist = _levenshtein_distance(a_low, b_low)
    return 1.0 - dist / max_len


class ColumnAmbiguityError(ValueError):
    """Raised when multiple columns match a hint with similarity >= 0.8."""
    def __init__(self, hint: str, candidates: list[dict]):
        self.hint = hint
        self.candidates = candidates
        super().__init__(
            f"Column mapping ambiguity for hint '{hint}': "
            f"multiple candidates found: {candidates}"
        )


def pick_col(df: pd.DataFrame, candidates: list[str], threshold: float = 0.7) -> Optional[str]:
    """
    Finds the best matching column via Levenshtein distance.
    
    PRD §7.3: similarity threshold 0.7; if multiple columns score >= 0.8,
    raise ColumnAmbiguityError so the Agent can prompt the user.
    """
    best_col: Optional[str] = None
    best_score: float = 0.0
    high_matches: list[dict] = []  # columns with score >= 0.8

    for col_name in df.columns:
        col_str = str(col_name).strip()
        for cand in candidates:
            score = _similarity(col_str, cand)
            # Also check substring containment as a bonus signal
            if cand.lower() in col_str.lower():
                score = max(score, 0.85)
            
            if score >= 0.8:
                high_matches.append({"column": col_str, "hint": cand, "similarity": round(score, 3)})
            if score > best_score:
                best_score = score
                best_col = col_str

    # Ambiguity check: if 2+ distinct columns scored >= 0.8, raise error
    unique_cols = {m["column"] for m in high_matches}
    if len(unique_cols) > 1:
        raise ColumnAmbiguityError(
            hint=candidates[0] if candidates else "",
            candidates=high_matches,
        )

    if best_score >= threshold:
        return best_col
    return None


def parse_spreadsheet(file_path: str, sheet_name: Optional[str] = None) -> List[Dict[str, Any]]:
    """
    Parses a spreadsheet into normalized list of dicts.
    Handles xlsx, xls, csv, and fills merged cells dynamically to prevent NaNs.
    """
    ext = file_path.rsplit(".", 1)[-1].lower()
    
    if ext == "csv":
        df = pd.read_csv(file_path)
    elif ext in ["xls"]:
        df = pd.read_excel(file_path, sheet_name=sheet_name or 0)
    elif ext in ["xlsx", "xlsm"]:
        # openpyxl preprocessing to handle merged cells
        wb = load_workbook(file_path, data_only=True)
        sheet = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
        if not sheet:
            raise ValueError(f"Could not load sheet {sheet_name}")

        _fill_merged_cells(sheet)
        
        # Convert fixed sheet back to Pandas DF via generator
        data = list(sheet.values)  # type: ignore
        if not data:
            raise ValueError("Empty excel document")

        # Heuristic header row detection
        best_row_idx = 0
        best_header_score = 0
        
        # known targets to score rows
        targets = {"物料", "品名", "名称", "name", "item", "数量", "qty", "quantity", "目标数量", "target", "数目", "单价", "price", "unit", "类别", "类型", "category", "type"}
        
        for i in range(min(15, len(data))):
            row = data[i]
            score = 0
            for cell in row:
                c_str = str(cell).strip().lower()
                if c_str in targets or any(kw in c_str for kw in targets):
                    score += 1
            if score > best_header_score:
                best_header_score = score
                best_row_idx = i
                
        cols = data[best_row_idx]
        cols = [f"col_{j}" if c is None else str(c).strip() for j, c in enumerate(cols)]
        
        df = pd.DataFrame(data[best_row_idx+1:], columns=cols)
    else:
        raise ValueError(f"Unsupported spreadsheet extension: {ext}")

    # Normalize columns explicitly (to handle unnamed correctly if duplicate)
    s = pd.Series(df.columns)
    df.columns = s.where(~s.duplicated(), s + "_" + s.groupby(s).cumcount().astype(str))

    # Matching column dictionaries
    col_name = pick_col(df, ["物料", "品名", "名称", "name", "item"])
    col_qty = pick_col(df, ["数量", "qty", "quantity", "目标数量", "target", "数目"])
    col_price = pick_col(df, ["单价", "price", "unit"])
    col_cat = pick_col(df, ["类别", "类型", "category", "type"])

    if not col_name:
        raise ValueError("Missing 'name' column in excel document. Checked variants: [物料, 品名, 名称]")

    items: List[Dict[str, Any]] = []
    
    for _, row in df.iterrows():
        name = str(row.get(col_name, "")).strip()
        if not name or name.lower() in ("nan", "none", "<na>"):
            continue
            
        target_qty = row.get(col_qty, 0) if col_qty else 0
        unit_price = row.get(col_price, "") if col_price else ""
        category = str(row.get(col_cat, "")).strip() if col_cat else "field_photo"
        
        try:
            qty_int = int(float(target_qty)) if target_qty not in ("", None, "nan") and not pd.isna(target_qty) else 0
            if qty_int < 0:
                qty_int = 0
        except Exception:
            qty_int = 0

        items.append(
            {
                "name": name,
                "category": category or "field_photo",
                "target_qty": qty_int,
                "unit_price": str(unit_price) if unit_price is not None and not pd.isna(unit_price) else "",
                "sort_order": len(items) + 1,
            }
        )

    return items
